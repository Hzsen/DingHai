import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def normalize_code(series):
    return series.astype(str).str.extract(r"(\d+)")[0].fillna("").str.zfill(6)


def extract_rank_dates(path):
    match = re.search(r"stock_rank_change_(20\d{6})_(20\d{6})", path)
    if match:
        return match.group(1), match.group(2)
    return None, None


def extract_lianban_date(path):
    match = re.search(r"封板(\d{8})", path)
    if match:
        return match.group(1)
    return None


def get_fill_map():
    return {
        1: "00B050",  # green
        2: "0070C0",  # blue
        3: "C0504D",  # onion red
        4: "C47F3A",  # ochre
        5: "FFFF00",  # yellow
        6: "FF0000",  # red
        7: "808080",  # gray
        8: "F4B084",  # peach
    }


def apply_colors(
    output_path, code_col_name="代码", name_col_name="名称", lianban_col_name="连板天"
):
    wb = load_workbook(output_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if code_col_name not in header or lianban_col_name not in header:
        raise ValueError("输出文件缺少必要列：代码 或 连板天")
    code_idx = header.index(code_col_name) + 1
    name_idx = header.index(name_col_name) + 1 if name_col_name in header else None
    lianban_idx = header.index(lianban_col_name) + 1

    fills = {
        k: PatternFill(fill_type="solid", start_color=v, end_color=v)
        for k, v in get_fill_map().items()
    }

    for row in range(2, ws.max_row + 1):
        lianban_cell = ws.cell(row=row, column=lianban_idx)
        code_cell = ws.cell(row=row, column=code_idx)
        name_cell = ws.cell(row=row, column=name_idx) if name_idx else None
        try:
            value = int(float(lianban_cell.value))
        except Exception:
            continue
        if value >= 8:
            code_cell.fill = fills[8]
            if name_cell:
                name_cell.fill = fills[8]
        elif value in fills:
            code_cell.fill = fills[value]
            if name_cell:
                name_cell.fill = fills[value]

    wb.save(output_path)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="合并涨幅排名与连板天数据")
    parser.add_argument(
        "--rank",
        default="data/stock_rank_change_20260203_20260204.xlsx",
        help="merge0119.py 输出文件路径",
    )
    parser.add_argument(
        "--lianban",
        default="data/封板02022026.xlsx",
        help="color_dxb_lianban.py 输出文件路径",
    )
    args = parser.parse_args()

    rank_path = args.rank
    lianban_path = args.lianban

    rank_df = pd.read_excel(rank_path)
    lianban_df = pd.read_excel(lianban_path)

    if "代码" not in rank_df.columns:
        raise ValueError("排名文件缺少必要列：代码")
    if "代码" not in lianban_df.columns or "连板天" not in lianban_df.columns:
        raise ValueError("连板文件缺少必要列：代码 或 连板天")

    rank_df["代码"] = normalize_code(rank_df["代码"])
    lianban_df["代码"] = normalize_code(lianban_df["代码"])

    merged = pd.merge(
        rank_df,
        lianban_df[["代码", "连板天"]],
        on="代码",
        how="left",
    )
    merged["连板天"] = pd.to_numeric(merged["连板天"], errors="coerce").fillna(0)
    if "Delta" not in merged.columns:
        raise ValueError("合并结果缺少必要列：Delta")
    merged = merged.sort_values("Delta", ascending=False)

    lianban_date = extract_lianban_date(lianban_path)
    if lianban_date:
        output_name = f"连板标记合集{lianban_date}.xlsx"
    else:
        output_name = "连板标记合集.xlsx"
    output_path = os.path.join("data", output_name)

    merged.to_excel(output_path, index=False)
    apply_colors(output_path)
    print(f"已保存: {output_path}")


if __name__ == "__main__":
    main()
