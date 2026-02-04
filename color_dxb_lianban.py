import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def read_csv_with_encodings(path, encodings=None, **kwargs):
    encodings = encodings or ["utf-8-sig", "utf-8", "gb18030", "gbk"]
    last_error = None
    for encoding in encodings:
        try:
            return pd.read_csv(path, encoding=encoding, sep=None, engine="python", **kwargs)
        except Exception as exc:  # noqa: BLE001 - best-effort encoding fallback
            last_error = exc
    raise last_error


def read_data(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".csv", ".txt"]:
        return read_csv_with_encodings(path)
    try:
        return pd.read_excel(path, engine="xlrd")
    except Exception:
        try:
            return pd.read_excel(path, engine="openpyxl")
        except Exception:
            return read_csv_with_encodings(path)


def normalize_columns(df):
    df.columns = [re.sub(r"\s+", "", str(c)) for c in df.columns]
    code_col = next((c for c in df.columns if "代码" in c), None)
    lianban_col = next((c for c in df.columns if "连板天" in c), None)
    if not code_col or not lianban_col:
        raise ValueError("未找到必要列：代码 或 连板天")
    rename_map = {}
    if code_col != "代码":
        rename_map[code_col] = "代码"
    if lianban_col != "连板天":
        rename_map[lianban_col] = "连板天"
    df = df.rename(columns=rename_map)
    df["代码"] = df["代码"].astype(str).str.extract(r"(\d+)")[0].fillna("").str.zfill(6)
    return df


def extract_mmddyyyy(path):
    match = re.search(r"(20\d{6})", path)
    if match:
        yyyymmdd = match.group(1)
        return f"{yyyymmdd[4:6]}{yyyymmdd[6:8]}{yyyymmdd[0:4]}"
    return datetime.now().strftime("%m%d%Y")


def get_fill_map():
    return {
        1: "00B050",  # green
        2: "0070C0",  # blue
        3: "C0504D",  # onion red
        4: "C47F3A",  # ochre
        5: "FFFF00",  # yellow
        6: "FF0000",  # red
        7: "808080",  # gray
        8: "F4B084",  # peach
    }


def apply_colors(
    output_path, code_col_name="代码", name_col_name="名称", lianban_col_name="连板天"
):
    wb = load_workbook(output_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if code_col_name not in header or lianban_col_name not in header:
        raise ValueError("输出文件缺少必要列：代码 或 连板天")
    code_idx = header.index(code_col_name) + 1
    name_idx = header.index(name_col_name) + 1 if name_col_name in header else None
    lianban_idx = header.index(lianban_col_name) + 1

    fills = {
        k: PatternFill(fill_type="solid", start_color=v, end_color=v)
        for k, v in get_fill_map().items()
    }

    for row in range(2, ws.max_row + 1):
        lianban_cell = ws.cell(row=row, column=lianban_idx)
        code_cell = ws.cell(row=row, column=code_idx)
        name_cell = ws.cell(row=row, column=name_idx) if name_idx else None
        try:
            value = int(float(lianban_cell.value))
        except Exception:
            continue
        if value >= 8:
            code_cell.fill = fills[8]
            if name_cell:
                name_cell.fill = fills[8]
        elif value in fills:
            code_cell.fill = fills[value]
            if name_cell:
                name_cell.fill = fills[value]

    wb.save(output_path)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="按连板天为代码/名称上色")
    parser.add_argument(
        "input_path",
        nargs="?",
        default="data/短线宝20260202.xls",
        help="输入文件路径，如 data/短线宝20260202.xls",
    )
    args = parser.parse_args()
    input_path = args.input_path
    df = normalize_columns(read_data(input_path))
    mmddyyyy = extract_mmddyyyy(input_path)
    output_path = os.path.join("data", f"封板{mmddyyyy}.xlsx")
    df.to_excel(output_path, index=False)
    apply_colors(output_path)
    print(f"已保存: {output_path}")


if __name__ == "__main__":
    main()
